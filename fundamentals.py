"""Company fundamentals data pipeline (market cap, EV/EBITDA, revenue, etc.)
for the Excel export feature.

Three external data sources, routed by ticker exchange suffix:
  - KRX (data.krx.co.kr Open API): market cap / price / shares outstanding for
    Korean-listed tickers (.KS / .KQ).
  - DART (opendart.fss.or.kr Open API): disclosed financial statements
    (revenue, operating income, debt, cash) for Korean-listed tickers.
    DART only has historical actual figures, never forward estimates.
  - FMP (financialmodelingprep.com): fundamentals for every other ticker,
    and the sole source of FY25/FY26 forward analyst estimates for ANY ticker
    (KRX+DART never provide forward estimates).

Every external call is individually wrapped so that a missing API key, an
unsupported ticker, or a provider outage degrades only the affected field(s)
to `None` — it never raises out of `fetch_fundamentals_for_ticker()` and never
aborts a run.
"""

import io
import logging
import os
import time
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone

import requests

logger = logging.getLogger(__name__)

DART_API_KEY = os.getenv("DART_API_KEY", "")
KRX_API_KEY = os.getenv("KRX_API_KEY", "")
FMP_API_KEY = os.getenv("FMP_API_KEY", "")

FUNDAMENTALS_CACHE_TTL_SECONDS = int(os.getenv("FUNDAMENTALS_CACHE_TTL_SECONDS", "21600"))

YAHOO_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

DART_BASE = "https://opendart.fss.or.kr/api"
KRX_BASE = "http://data-dbg.krx.co.kr/svc/apis"
FMP_BASE = "https://financialmodelingprep.com/api/v3"

# ─── Excel column spec (order matters — exactly as requested) ─────────────────
# (key, Korean header label, number_format, is_usd)
EXCEL_COLUMNS = [
    ("name",                   "회사명",                                    None,             False),
    ("market_cap_usd",         "시총",                                      "#,##0",          True),
    ("net_debt_usd",           "순차입",                                    "#,##0",          True),
    ("change_pct_1d",          "전일비 등락율",                             '0.00"%"',        False),
    ("change_pct_2d",          "2일전등락율",                               '0.00"%"',        False),
    ("change_pct_3d",          "3일전등락율",                               '0.00"%"',        False),
    ("mktcap_52w_low_usd",     "52주최저 시총",                             "#,##0",          True),
    ("mktcap_52w_high_usd",    "52주 최고시총",                             "#,##0",          True),
    ("mktcap_52w_trend",       "52주 시총추이",                             None,             False),
    ("pct_of_52w_high",        "52주 최고시총비현재가",                     '0.00"%"',        False),
    ("ev_ebitda_fy25",         "EV/EBITDA(현재EV/25년말 EBITDA)",           "0.00",           False),
    ("ev_ebitda_fy26",         "EV/EBITDA(현재/26년말 EBITDA)",             "0.00",           False),
    ("price_local",            "주가(현지통화)",                            "#,##0.00",       False),
    ("revenue_fy25_usd",       "매출(FY25)",                                "#,##0",          True),
    ("revenue_fy26_usd",       "매출(FY26)",                                "#,##0",          True),
    ("operating_margin_fy25",  "영업이익율(FY25)",                          '0.00"%"',        False),
    ("operating_margin_fy26",  "영업이익율(FY26)",                          '0.00"%"',        False),
    ("ebitda_fy25_usd",        "EBITDA(FY25)",                              "#,##0",          True),
    ("ebitda_fy26_usd",        "EBITDA(FY26)",                              "#,##0",          True),
    ("ev_usd",                 "EV",                                        "#,##0",          True),
]

_SPARK_CHARS = "▁▂▃▄▅▆▇█"

# ─── Small in-process TTL caches ───────────────────────────────────────────────
_fundamentals_cache = {}   # ticker -> (fetched_at_epoch, record)
_fx_cache = {}             # currency -> (fetched_at_epoch, rate)


def ticker_source(ticker):
    """Route a ticker to its primary fundamentals data source."""
    t = (ticker or "").upper()
    if t.endswith((".KS", ".KQ")):
        return "KR"
    return "OTHER"


def _yahoo_chart(ticker, range_str="1y", interval="1d"):
    """Minimal standalone Yahoo Finance chart fetch (mirrors app.py's
    fetch_single_ticker but kept local to avoid a circular import with app.py).
    Returns the raw `result` dict (meta/timestamp/indicators) or None.
    """
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
        resp = requests.get(
            url, params={"range": range_str, "interval": interval},
            headers=YAHOO_HEADERS, timeout=10,
        )
        if resp.status_code != 200:
            return None
        data = resp.json()
        result = data["chart"]["result"][0]
        return result
    except Exception as e:
        logger.warning(f"Yahoo chart fetch failed for {ticker}: {e}")
        return None


def get_fx_rate(currency_code):
    """USD per 1 unit of currency_code. Returns 1.0 for USD or on failure
    (better to leave a value unconverted-but-present than to blank it)."""
    if not currency_code or currency_code.upper() == "USD":
        return 1.0

    now = time.time()
    cached = _fx_cache.get(currency_code)
    if cached and (now - cached[0]) < 3600:
        return cached[1]

    try:
        result = _yahoo_chart(f"{currency_code.upper()}USD=X", range_str="5d")
        if result is None:
            # Fallback pair direction
            result = _yahoo_chart(f"{currency_code.upper()}=X", range_str="5d")
            if result is None:
                return cached[1] if cached else 1.0
            closes = [c for c in result["indicators"]["quote"][0]["close"] if c is not None]
            if not closes:
                return cached[1] if cached else 1.0
            rate = 1.0 / closes[-1]
        else:
            closes = [c for c in result["indicators"]["quote"][0]["close"] if c is not None]
            if not closes:
                return cached[1] if cached else 1.0
            rate = closes[-1]
        _fx_cache[currency_code] = (now, rate)
        return rate
    except Exception as e:
        logger.warning(f"FX rate fetch failed for {currency_code}: {e}")
        return cached[1] if cached else 1.0


def fetch_recent_changes(ticker, target_date=None, tz_hours=0, n=3):
    """Return up to n most recent daily % changes, most recent first:
    [chg_1d, chg_2d, chg_3d]. Missing values are None."""
    try:
        result = _yahoo_chart(ticker, range_str="3mo")
        if result is None:
            return [None] * n
        closes = result["indicators"]["quote"][0]["close"]
        timestamps = result["timestamp"]
        market_tz = timezone(timedelta(hours=tz_hours))
        valid = [(ts, c) for ts, c in zip(timestamps, closes) if c is not None]
        seen = {}
        for ts, c in valid:
            local_date = datetime.fromtimestamp(ts, market_tz).date()
            seen[local_date] = (ts, c)
        valid = sorted(seen.values())
        if target_date:
            valid = [(ts, c) for ts, c in valid
                     if datetime.fromtimestamp(ts, market_tz).date() <= target_date]
        closes_only = [c for _, c in valid]
        changes = []
        for i in range(n):
            idx_last = len(closes_only) - 1 - i
            idx_prev = idx_last - 1
            if idx_last < 0 or idx_prev < 0:
                changes.append(None)
                continue
            prev, last = closes_only[idx_prev], closes_only[idx_last]
            if not prev:
                changes.append(None)
                continue
            changes.append(round(((last - prev) / prev) * 100, 2))
        return changes
    except Exception as e:
        logger.warning(f"fetch_recent_changes failed for {ticker}: {e}")
        return [None] * n


def _sparkline(values):
    vals = [v for v in values if v is not None]
    if len(vals) < 2:
        return None
    lo, hi = min(vals), max(vals)
    span = hi - lo
    out = []
    for v in values:
        if v is None:
            out.append(" ")
            continue
        if span == 0:
            out.append(_SPARK_CHARS[0])
            continue
        idx = int((v - lo) / span * (len(_SPARK_CHARS) - 1))
        out.append(_SPARK_CHARS[idx])
    return "".join(out)


# ─── DART (Korea disclosed financial statements) ──────────────────────────────

_dart_corp_code_cache_file = "dart_corp_codes.json"
_dart_corp_code_map = None  # stock_code (6-digit) -> corp_code


def _dart_corp_code_lookup(krx_code_6digit):
    """Map a 6-digit KRX stock code to a DART corp_code, downloading and
    caching the full corpCode.xml mapping on first use."""
    global _dart_corp_code_map
    if not DART_API_KEY:
        return None
    if _dart_corp_code_map is None:
        _dart_corp_code_map = _load_or_fetch_dart_corp_codes()
    return _dart_corp_code_map.get(krx_code_6digit)


def _load_or_fetch_dart_corp_codes():
    import json as _json
    if os.path.exists(_dart_corp_code_cache_file):
        try:
            with open(_dart_corp_code_cache_file, "r", encoding="utf-8") as f:
                return _json.load(f)
        except Exception:
            pass
    mapping = {}
    try:
        resp = requests.get(f"{DART_BASE}/corpCode.xml", params={"crtfc_key": DART_API_KEY}, timeout=30)
        resp.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
            with zf.open("CORPCODE.xml") as xf:
                tree = ET.parse(xf)
        for item in tree.getroot().findall("list"):
            stock_code = (item.findtext("stock_code") or "").strip()
            corp_code = (item.findtext("corp_code") or "").strip()
            if stock_code:
                mapping[stock_code] = corp_code
        try:
            with open(_dart_corp_code_cache_file, "w", encoding="utf-8") as f:
                _json.dump(mapping, f)
        except Exception:
            pass
    except Exception as e:
        logger.warning(f"DART corpCode.xml fetch failed: {e}")
    return mapping


# Known label variants for each line item (DART account_nm isn't perfectly
# standardized across filers).
_DART_LABELS = {
    "revenue": ["매출액", "수익(매출액)", "영업수익"],
    "operating_income": ["영업이익", "영업이익(손실)"],
    "total_debt": ["단기차입금", "장기차입금", "사채", "유동성장기부채"],
    "cash": ["현금및현금성자산", "단기금융상품"],
}


def _dart_financials(corp_code, year, reprt_code="11011"):
    """Fetch one fiscal year's consolidated (CFS, fallback OFS) financial
    statement line items from DART. Returns dict with revenue/operating_income
    /total_debt/cash in KRW, any of which may be None if not found."""
    if not DART_API_KEY or not corp_code:
        return {}
    try:
        resp = requests.get(
            f"{DART_BASE}/fnlttSinglAcntAll.json",
            params={
                "crtfc_key": DART_API_KEY, "corp_code": corp_code,
                "bsns_year": str(year), "reprt_code": reprt_code, "fs_div": "CFS",
            },
            timeout=15,
        )
        data = resp.json()
        rows = data.get("list", [])
        if not rows:
            resp = requests.get(
                f"{DART_BASE}/fnlttSinglAcntAll.json",
                params={
                    "crtfc_key": DART_API_KEY, "corp_code": corp_code,
                    "bsns_year": str(year), "reprt_code": reprt_code, "fs_div": "OFS",
                },
                timeout=15,
            )
            rows = resp.json().get("list", [])

        result = {"revenue": None, "operating_income": None, "total_debt": 0.0, "cash": 0.0}
        found_debt = False
        found_cash = False
        for row in rows:
            name = (row.get("account_nm") or "").strip()
            try:
                amount = float((row.get("thstrm_amount") or "0").replace(",", ""))
            except ValueError:
                continue
            if name in _DART_LABELS["revenue"] and result["revenue"] is None:
                result["revenue"] = amount
            elif name in _DART_LABELS["operating_income"] and result["operating_income"] is None:
                result["operating_income"] = amount
            elif name in _DART_LABELS["total_debt"]:
                result["total_debt"] += amount
                found_debt = True
            elif name in _DART_LABELS["cash"]:
                result["cash"] += amount
                found_cash = True
        if not found_debt:
            result["total_debt"] = None
        if not found_cash:
            result["cash"] = None
        return result
    except Exception as e:
        logger.warning(f"DART financials fetch failed for corp_code={corp_code}: {e}")
        return {}


def _compute_net_debt(dart_financials):
    total_debt = dart_financials.get("total_debt")
    cash = dart_financials.get("cash")
    if total_debt is None or cash is None:
        return None
    return total_debt - cash


# ─── KRX (Korea market data) ───────────────────────────────────────────────────

def _krx_headers():
    return {"AUTH_KEY": KRX_API_KEY} if KRX_API_KEY else {}


def _krx_snapshot(krx_code_6digit, bas_dd):
    """Single-day cross-sectional row for one KR stock: close price, market
    cap, listed shares. bas_dd format: YYYYMMDD."""
    if not KRX_API_KEY:
        return None
    try:
        resp = requests.get(
            f"{KRX_BASE}/sto/stk_bydd_trd",
            params={"basDd": bas_dd},
            headers=_krx_headers(), timeout=15,
        )
        rows = resp.json().get("OutBlock_1", [])
        for row in rows:
            if row.get("ISU_CD", "").endswith(krx_code_6digit) or row.get("ISU_SRT_CD") == krx_code_6digit:
                return {
                    "close": _to_float(row.get("TDD_CLSPRC")),
                    "market_cap": _to_float(row.get("MKTCAP")),
                    "shares": _to_float(row.get("LIST_SHRS")),
                }
        return None
    except Exception as e:
        logger.warning(f"KRX snapshot fetch failed for {krx_code_6digit}@{bas_dd}: {e}")
        return None


def _to_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def _krx_52w_series(krx_code_6digit, target_date):
    """~12 monthly snapshots over the trailing 52 weeks. Returns list of
    market-cap values (may contain None gaps)."""
    if not KRX_API_KEY:
        return []
    series = []
    for months_back in range(11, -1, -1):
        snap_date = target_date - timedelta(days=30 * months_back)
        bas_dd = snap_date.strftime("%Y%m%d")
        row = _krx_snapshot(krx_code_6digit, bas_dd)
        series.append(row["market_cap"] if row else None)
    return series


# ─── FMP (non-Korean fundamentals + universal forward estimates) ─────────────

def _fmp_get(path, **params):
    if not FMP_API_KEY:
        return None
    try:
        params["apikey"] = FMP_API_KEY
        resp = requests.get(f"{FMP_BASE}/{path}", params=params, timeout=15)
        if resp.status_code != 200:
            return None
        return resp.json()
    except Exception as e:
        logger.warning(f"FMP fetch failed for {path}: {e}")
        return None


def _fmp_profile(symbol):
    data = _fmp_get(f"profile/{symbol}")
    if not data:
        return {}
    row = data[0] if isinstance(data, list) and data else {}
    return {
        "market_cap": row.get("mktCap"),
        "price": row.get("price"),
        "currency": row.get("currency"),
    }


def _fmp_enterprise_values(symbol, period="quarter", limit=5):
    data = _fmp_get(f"enterprise-values/{symbol}", period=period, limit=limit)
    return data if isinstance(data, list) else []


def _fmp_analyst_estimates(symbol, period="annual", limit=10):
    data = _fmp_get(f"analyst-estimates/{symbol}", period=period, limit=limit)
    return data if isinstance(data, list) else []


def _fmp_estimate_for_year(estimates, year):
    for row in estimates:
        date_str = row.get("date", "")
        if date_str.startswith(str(year)):
            return row
    return None


# ─── Main orchestrator ─────────────────────────────────────────────────────────

def fetch_fundamentals_for_ticker(ticker, meta=None, target_date=None):
    """Cached wrapper around _fetch_fundamentals_uncached — avoids re-hitting
    paid external APIs on repeated manual runs within FUNDAMENTALS_CACHE_TTL_SECONDS."""
    now = time.time()
    cached = _fundamentals_cache.get(ticker)
    if cached and (now - cached[0]) < FUNDAMENTALS_CACHE_TTL_SECONDS:
        return cached[1]
    record = _fetch_fundamentals_uncached(ticker, meta=meta, target_date=target_date)
    _fundamentals_cache[ticker] = (now, record)
    return record


def _fetch_fundamentals_uncached(ticker, meta=None, target_date=None):
    """Build one complete fundamentals record (all EXCEL_COLUMNS keys present,
    value None where unavailable). Never raises."""
    meta = meta or {}
    record = {key: None for key, _, _, _ in EXCEL_COLUMNS}
    record["name"] = meta.get("name") or ticker
    source = ticker_source(ticker)

    try:
        tz_hours = 9 if source == "KR" else 0
        changes = fetch_recent_changes(ticker, target_date=target_date, tz_hours=tz_hours)
        record["change_pct_1d"], record["change_pct_2d"], record["change_pct_3d"] = (
            changes + [None] * 3
        )[:3]
    except Exception as e:
        logger.warning(f"Recent changes failed for {ticker}: {e}")

    try:
        if source == "KR":
            _fill_kr_fundamentals(ticker, record, target_date)
        else:
            _fill_other_fundamentals(ticker, record, target_date)
    except Exception as e:
        logger.warning(f"Fundamentals fetch failed for {ticker}: {e}")
        record["error"] = str(e)

    try:
        _fill_forward_estimates(ticker, record)
    except Exception as e:
        logger.warning(f"Forward estimates failed for {ticker}: {e}")

    try:
        _finalize_derived_fields(record)
    except Exception as e:
        logger.warning(f"Derived fields computation failed for {ticker}: {e}")

    return record


def _fill_kr_fundamentals(ticker, record, target_date):
    krx_code = ticker.split(".")[0]
    today = target_date or datetime.now().date()

    snap = _krx_snapshot(krx_code, today.strftime("%Y%m%d"))
    fx = get_fx_rate("KRW")
    if snap:
        record["price_local"] = snap.get("close")
        if snap.get("market_cap") is not None:
            record["market_cap_usd"] = snap["market_cap"] * fx

    series = _krx_52w_series(krx_code, today)
    valid_series = [v for v in series if v is not None]
    if valid_series:
        record["mktcap_52w_low_usd"] = min(valid_series) * fx
        record["mktcap_52w_high_usd"] = max(valid_series) * fx
        record["mktcap_52w_trend"] = _sparkline(series)

    corp_code = _dart_corp_code_lookup(krx_code)
    if corp_code:
        fin = _dart_financials(corp_code, today.year - 1)  # latest confirmed annual report
        net_debt_krw = _compute_net_debt(fin)
        if net_debt_krw is not None:
            record["net_debt_usd"] = net_debt_krw * fx


def _fill_other_fundamentals(ticker, record, target_date):
    profile = _fmp_profile(ticker)
    currency = profile.get("currency") or "USD"
    fx = get_fx_rate(currency)

    if profile.get("price") is not None:
        record["price_local"] = profile["price"]
    if profile.get("market_cap") is not None:
        record["market_cap_usd"] = profile["market_cap"] * fx

    ev_series = _fmp_enterprise_values(ticker, period="quarter", limit=5)
    if ev_series:
        latest = ev_series[0]
        market_caps = [row.get("marketCapitalization") for row in ev_series if row.get("marketCapitalization") is not None]
        if market_caps:
            record["mktcap_52w_low_usd"] = min(market_caps) * fx
            record["mktcap_52w_high_usd"] = max(market_caps) * fx
            record["mktcap_52w_trend"] = _sparkline(list(reversed(market_caps)))
        net_debt = latest.get("netDebt") or latest.get("addTotalDebt")
        if net_debt is not None:
            record["net_debt_usd"] = net_debt * fx
        ev = latest.get("enterpriseValue")
        if ev is not None:
            record["ev_usd"] = ev * fx


def _fill_forward_estimates(ticker, record):
    """FY25/FY26 estimates — FMP only, for ANY ticker regardless of KR/OTHER."""
    estimates = _fmp_analyst_estimates(ticker, period="annual", limit=10)
    if not estimates:
        return
    fx = None  # analyst-estimates values are typically already in the company's reporting currency
    for fy, rev_key, opinc_key, ebitda_key in (
        (25, "revenue_fy25_usd", "operating_margin_fy25", "ebitda_fy25_usd"),
        (26, "revenue_fy26_usd", "operating_margin_fy26", "ebitda_fy26_usd"),
    ):
        year = 2000 + fy
        row = _fmp_estimate_for_year(estimates, year)
        if not row:
            continue
        if fx is None:
            profile = _fmp_profile(ticker) if ticker_source(ticker) != "KR" else {}
            currency = profile.get("currency") or ("KRW" if ticker_source(ticker) == "KR" else "USD")
            fx = get_fx_rate(currency)
        revenue = row.get("estimatedRevenueAvg")
        ebitda = row.get("estimatedEbitdaAvg")
        op_income = row.get("estimatedOperatingIncomeAvg")
        if revenue is not None:
            record[rev_key] = revenue * fx
        if ebitda is not None:
            record[ebitda_key] = ebitda * fx
        if revenue and op_income is not None:
            record[opinc_key] = (op_income / revenue) * 100


def _finalize_derived_fields(record):
    market_cap = record.get("market_cap_usd")
    net_debt = record.get("net_debt_usd")
    if record.get("ev_usd") is None and market_cap is not None and net_debt is not None:
        record["ev_usd"] = market_cap + net_debt

    high = record.get("mktcap_52w_high_usd")
    if market_cap is not None and high:
        record["pct_of_52w_high"] = (market_cap / high) * 100

    ev = record.get("ev_usd")
    ebitda_fy25 = record.get("ebitda_fy25_usd")
    ebitda_fy26 = record.get("ebitda_fy26_usd")
    if ev is not None and ebitda_fy25:
        record["ev_ebitda_fy25"] = ev / ebitda_fy25
    if ev is not None and ebitda_fy26:
        record["ev_ebitda_fy26"] = ev / ebitda_fy26


# ─── Excel workbook builder ─────────────────────────────────────────────────────

def build_fundamentals_workbook(fundamentals, ticker_objects):
    """fundamentals: dict ticker -> record (from fetch_fundamentals_for_ticker).
    ticker_objects: list of {ticker, category, name} defining row order.
    Returns an openpyxl.Workbook."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Fundamentals"

    headers = [label for _, label, _, _ in EXCEL_COLUMNS]
    ws.append(headers)
    ws.freeze_panes = "A2"

    for t in ticker_objects:
        record = fundamentals.get(t["ticker"], {})
        row = [record.get(key) for key, _, _, _ in EXCEL_COLUMNS]
        ws.append(row)

    last_col_letter = ws.cell(row=1, column=len(EXCEL_COLUMNS)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    for col_idx, (_, _, number_format, _) in enumerate(EXCEL_COLUMNS, start=1):
        if number_format is None:
            continue
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        for row_idx in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = number_format

    return wb
